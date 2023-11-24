from django.shortcuts import render, redirect, get_object_or_404
from django.forms.models import model_to_dict
from django.contrib import messages
from .models import *
from ..Usuarios.models import Trabajadores
from django.http import JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
import os
import io
import datetime
from django.conf import settings

# Create your views here.

#Importamos Librerias
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF, HTMLMixin

#Home
def bodega(request, codigo=None):
    listarBodega = Bodega.objects.all()
    listarTrabajadores = Trabajadores.objects.all()
    listarProveedores = Proveedores.objects.all()
    listarProductos = Productos.objects.all()
    
    return render(request, "bodega/panel.html", {'Bodega':listarBodega, 'Trabajadores': listarTrabajadores,'Proveedores': listarProveedores, 'Productos': listarProductos, 'Codigo':codigo})

def listadoBodega(request):
    listarBodega = Bodega.objects.all()
    
    return render(request, "bodega/listadoBodega.html", {'Bodega':listarBodega})

def registrarBodega(request):
    if request.method == 'POST':
        codigoEntrega = request.POST['primaryKey']
        proveedor_codigo = request.POST['proveedor']
        proveedor = Proveedores.objects.get(codigoProveedor=int(proveedor_codigo))
        producto_codigo = request.POST['producto']
        producto = Productos.objects.get(codigoProducto=int(producto_codigo))
        cantidad = request.POST['cantidad']
        pago = request.POST['pago']
        fecha = request.POST['fecha']
        trabajador_cedula = request.POST['trabajador']
        trabajador = Trabajadores.objects.get(cedula=int(trabajador_cedula))

        bodega = Bodega.objects.create(
            codigoEntrega = codigoEntrega,
            proveedor = proveedor,
            producto = producto, 
            cantidad = cantidad,
            pago = pago,
            fecha = fecha,
            trabajador = trabajador
        )
        
        return redirect('/bodega/')    
    else:
        return redirect('/')
    
def obtener_bodega(request, codigo):
    if request.method == 'GET':
        #Obtiene el elemento de la tabla Trabajadores que coincida con la cedula
        bodega = get_object_or_404(Bodega, codigoEntrega=codigo)
        data = {
            'proveedor': bodega.proveedor.codigoProveedor,
            'cantidad': bodega.cantidad,
            'pago': bodega.pago,
            'trabajadorId': bodega.trabajador_id,
            'producto': bodega.producto.codigoProducto,
            'fecha': bodega.fecha,
        }
        return JsonResponse({'data': data})
    
def editarBodega(request, codigo):

    registro = Bodega.objects.get(codigoEntrega=int(codigo))

    if request.method == 'POST':

        #Obtener los datos del formulario

        proveedor_codigo = request.POST['proveedor']
        proveedor = Proveedores.objects.get(codigoProveedor=int(proveedor_codigo))
        producto_codigo = request.POST['producto']
        producto = Productos.objects.get(codigoProducto=int(producto_codigo))
        cantidad = request.POST['cantidad']
        pago = request.POST['pago']
        fecha = request.POST['fecha']
        trabajador_cedula = request.POST['trabajador']
        trabajador = Trabajadores.objects.get(cedula=trabajador_cedula)

        registro.proveedor = proveedor
        registro.producto = producto
        registro.cantidad = cantidad
        registro.pago = pago
        registro.fecha = fecha
        registro.trabajador = trabajador

        registro.save()

        return redirect('/bodega/')
    
def eliminarBodega(request, codigo, url):

    #Elimina al trabajador

    registro = Bodega.objects.get(codigoEntrega=int(codigo))

    registro.delete()

    #Verificamos la url a la que redirige
    return redirect(f'/{url}/')

#Exportar Excel Bodega

def excelBodega(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Bodega.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:G1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Bodega'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Proveedor', 'Producto', 'Cantidad', 'Pago', 'Fecha', 'Trabajador']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoEntrega = registro.codigoEntrega
            proveedor = registro.proveedor.nombreProveedor
            producto = registro.producto.nombreProducto
            cantidad = registro.cantidad
            pago = registro.pago
            fecha = registro.fecha
            trabajador = registro.trabajador.nombreCompleto

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoEntrega).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=proveedor).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=producto).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=cantidad).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=pago).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=fecha).font = Font(name='Arial')
            sheet.cell(row=row_num, column=7, value=trabajador).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 7):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoBodega/')

def pdfBodega(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Bodega")

    pdf.tableBodega()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoBodega/')

class PDF(FPDF):
    # Método para crear el encabezado
    def encabezado(self, pagina):
        # Agregamos la imagen del logo izquierdo
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Astrocode.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 10, 8, 28)
        # Agregamos la imagen del logo derecho
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Logo Empresa.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 167, 8, 33)

        # Agregamos el título centrado y en negrita
        # Agregamos el título centrado y en negrita
        self.set_font('Arial', 'B', 24)
        self.set_text_color(88, 129, 87)
        self.cell(190, 15, 'Listado de ' + pagina, 0, 0, 'C')
        self.ln(20)

        # Agregamos el subtitulo centrado y en negrita
        # Obtener la fecha y hora actuales
        now = datetime.datetime.now()

        # Formatear la fecha y hora como una cadena con el formato deseado
        fecha_hora = now.strftime("%d/%m/%Y %H:%M:%S")
        
        # Agregar el subtitulo centrado y en negrita
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(0, -10, f"Exportado el: {fecha_hora}", 0, 0, 'C')
        self.ln(10)

        

    #Metodo para crear la tabla
    def tableBodega(self):
        #Registros Base de datos
        registros = Bodega.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [10, 37, 64, 91, 118, 145, 172]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Proveedor', 'Producto', 'Cantidad', 'Pago', 'Fecha', 'Trabajador']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoEntrega = str(registro.codigoEntrega)
            proveedor = registro.proveedor.nombreProveedor
            producto = registro.producto.nombreProducto
            cantidad = str(registro.cantidad)
            pago = str(registro.pago)
            fecha = str(registro.fecha)
            trabajador = registro.trabajador.nombreCompleto

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoEntrega, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, proveedor, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.cell(27, 11, producto, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, cantidad, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.cell(27, 11, pago, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.cell(27, 11, fecha, 1, 0, 'C', False, 0)

            self.set_x(x_coords[6])
            self.set_font('Arial', '', 6)
            self.cell(27, 11, trabajador, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableProveedores(self):
        #Registros Base de datos
        registros = Proveedores.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [10, 37, 64, 91, 118, 145]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Proveedor', 'Telefono', 'Direccion', 'Email', 'Ultima Entrega']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoProveedor = str(registro.codigoProveedor)
            nombreProveedor = registro.nombreProveedor
            telefono = str(registro.telefono)
            direccion = registro.direccion
            email = registro.email
            ultimaEntrega = str(registro.ultimaEntrega)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoProveedor, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, nombreProveedor, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.cell(27, 11, telefono, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.set_font('Arial', '', 7)
            self.cell(27, 11, direccion, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, email, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.set_font('Arial', '', 7)
            self.cell(27, 11, ultimaEntrega, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableCategorias(self):
        #Registros Base de datos
        registros = Categorias.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [80, 107]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Nombre']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoCategoria = str(registro.codigoCategorias)
            nombreCategoria = registro.nombreCategoria

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoCategoria, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, nombreCategoria, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableProductos(self):
        #Registros Base de datos
        registros = Productos.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [10, 37, 64, 91, 118, 145, 172]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Nombre', 'Descripcion', 'Precio Unidad', 'Stock', 'Proveedor', 'Categoria']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoProducto = str(registro.codigoProducto)
            nombreProducto = registro.nombreProducto
            descripcionProducto = registro.descripcionProducto
            precioUnidad = str(registro.precioUnidad)
            stock = str(registro.stock)
            proveedor = str(registro.proveedor.nombreProveedor)
            categoria = str(registro.categoria.nombreCategoria)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoProducto, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, nombreProducto, 1, 0, 'C', False, 0)
            
            self.set_x(x_coords[2])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, descripcionProducto, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, precioUnidad, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.cell(27, 11, stock, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.cell(27, 11, proveedor, 1, 0, 'C', False, 0)

            self.set_x(x_coords[6])
            self.cell(27, 11, categoria, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    # Método para crear el footer
    def footer(self):
        # Agregamos el texto del footer en gris
        self.set_y(-15)
        self.set_font('Arial', '', 8)
        self.set_text_color(127, 127, 127)
        self.cell(0, 10, 'Procol del Valle © Astrocode | Página ' + str(self.page_no()), 0, 0, 'C')

def verifyDirectories(directory, file):
    #Crea la carpeta donde se almacenan los datos
    directory_path = os.path.join(os.path.join(os.environ['USERPROFILE']), directory)
    procol_path = os.path.join(directory_path, 'Procol del Valle')

    if not os.path.exists(procol_path):
        os.makedirs(procol_path)
        excels_path = os.path.join(procol_path, 'Excels')
        pdf_path = os.path.join(procol_path, 'PDF')
        os.makedirs(excels_path)
        os.makedirs(pdf_path)

    if file == "Excel":
        excels_path = os.path.join(procol_path, 'Excels')

        if not os.path.exists(excels_path):
            os.makedirs(excels_path)

        return excels_path
    elif file == "Pdf":
        pdfs_path = os.path.join(procol_path, 'PDF')

        if not os.path.exists(pdfs_path):
            os.makedirs(pdfs_path)
        
        return pdfs_path
    






#HomeProductos
def productos(request, codigo=None):
    listarProductos = Productos.objects.all()
    listarCategorias = Categorias.objects.all()
    listarProveedores = Proveedores.objects.all()

    return render(request, "Productos/panel.html", {'Productos': listarProductos, 'Proveedores': listarProveedores, 'Categorias': listarCategorias, 'Codigo': codigo})

def listadoProductos(request):
    listarProductos = Productos.objects.all()
    listarProveedores = Proveedores.objects.all()
    return render(request, "Productos/listado.html", {'Productos': listarProductos, 'Proveedores': listarProveedores})

def registrarProducto(request):
    if request.method == 'POST':
        codigoProducto = request.POST['primaryKey']
        nombreProducto = request.POST['nombreProducto']
        descripcionProducto = request.POST['descripcionProducto']
        precioUnidad = request.POST['precio']
        stock = request.POST['stock']
        proveedor = request.POST['proveedor']
        categoria = request.POST['categoria']

        imagen = None
        if 'imageInput' in request.FILES:
            imagen = request.FILES['imageInput']

        proveedor = Proveedores.objects.get(codigoProveedor=proveedor)
        categoria = Categorias.objects.get(codigoCategorias=categoria)

        producto = Productos.objects.create(
            codigoProducto=codigoProducto,
            nombreProducto=nombreProducto,
            descripcionProducto=descripcionProducto,
            precioUnidad=precioUnidad,
            stock=stock,
            proveedor=proveedor,
            categoria=categoria,
            imagen=imagen,
        )

        return redirect('/Productos/')
    else:
        return redirect('/')


def obtener_producto(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Productos que coincida con el código
        produc = get_object_or_404(Productos, codigoProducto=codigo)
        data = {
            'nombreProducto': produc.nombreProducto,
            'descripcionProducto': produc.descripcionProducto,
            'precioUnidad': str(produc.precioUnidad),
            'stock': produc.stock,
            'proveedor': produc.proveedor.nombreProveedor,
            'categoria': produc.categoria.nombreCategoria,
            'imagen': produc.imagen.url if produc.imagen else None,
        }
        return JsonResponse({'data': data})


def editarProducto(request, codigo):
    producto = Productos.objects.get(codigoProducto=int(codigo))

    if request.method == 'POST':

        nombreProducto = request.POST['nombreProducto']
        descripcionProducto = request.POST['descripcionProducto']
        precioUnidad = request.POST['precio']
        stock = request.POST['stock']
        proveedor = request.POST['proveedor']
        categoria = request.POST['categoria']

        if 'imageInput' in request.FILES:
            imagen = request.FILES['imageInput']

            if producto.imagen != "":
                ruta_foto = "Files/" + str(producto.imagen)
                os.remove(ruta_foto)

            ruta_foto = "Files/Productos/" + str(imagen)
            with open(ruta_foto, 'wb') as f:
                for chunk in imagen.chunks():
                    f.write(chunk)

            producto.imagen = "/Productos/" + str(imagen)

        proveedor = Proveedores.objects.get(codigoProveedor=proveedor)
        categoria = Categorias.objects.get(codigoCategorias=categoria)

        producto.nombreProducto = nombreProducto
        producto.descripcionProducto = descripcionProducto
        producto.precioUnidad = precioUnidad
        producto.stock = stock
        producto.proveedor = proveedor
        producto.categoria = categoria

        producto.save()

        return redirect('/Productos/')

def eliminarProducto(request, codigo, url):
    producto = Productos.objects.get(codigoProducto=int(codigo))

    if producto.imagen != "":
        ruta_foto = "Files/" + str(producto.imagen)
        os.remove(ruta_foto)

    producto.delete()

    return redirect(f'/{url}/')

#Exportar Excel Bodega

def excelProductos(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Productos.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:G1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Productos'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Nombre', 'Descripcion', 'Precio Unidad', 'Stock', 'Proveedor', 'Categoria']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoProducto = registro.codigoProducto
            nombreProducto = registro.nombreProducto
            descripcionProducto = registro.descripcionProducto
            precioUnidad = registro.precioUnidad
            stock = registro.stock
            proveedor = registro.proveedor.nombreProveedor
            categoria = registro.categoria.nombreCategoria

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoProducto).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=nombreProducto).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=descripcionProducto).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=precioUnidad).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=stock).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=proveedor).font = Font(name='Arial')
            sheet.cell(row=row_num, column=7, value=categoria).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 8):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoProductos/')

def pdfProductos(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Productos")

    pdf.tableProductos()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoProductos/')


#Proveedores
def proveedores(request, codigo=None):
    listarProveedores = Proveedores.objects.all()
    
    return render(request, "Proveedores/panel.html", {'Proveedores': listarProveedores, 'Codigo': codigo})

def listadoProveedores(request):
    listarProveedores = Proveedores.objects.all()
    
    return render(request, "Proveedores/listado.html", {'Proveedores': listarProveedores})

def registrarProveedor(request):
    if request.method == 'POST':
        codigoProveedor = request.POST['primaryKey']
        nombreProveedor = request.POST['nombreProveedor']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']
        ultimaEntrega = request.POST['ultimaEntrega']

        proveedor = Proveedores.objects.create(
            codigoProveedor=codigoProveedor,
            nombreProveedor=nombreProveedor,
            telefono=telefono,
            direccion=direccion,
            email=email,
            ultimaEntrega=ultimaEntrega
        )

        return redirect('/Proveedores/')
    else:
        return redirect('/')

def obtener_proveedor(request, codigo):
    if request.method == 'GET':
        provee = get_object_or_404(Proveedores, codigoProveedor=codigo)
        data = {
            'codigoProveedor': provee.codigoProveedor,
            'nombreProveedor': provee.nombreProveedor,
            'telefono': provee.telefono,
            'direccion': provee.direccion,
            'email': provee.email,
            'ultimaEntrega': provee.ultimaEntrega,
        }
        return JsonResponse({'data': data})

def editarProveedor(request, codigo):
    
    proveedor = Proveedores.objects.get(codigoProveedor=int(codigo))

    if request.method == 'POST':

        nombreProveedor = request.POST['nombreProveedor']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']
        ultimaEntrega = request.POST['ultimaEntrega']


        proveedor.nombreProveedor = nombreProveedor
        proveedor.telefono = telefono
        proveedor.direccion = direccion
        proveedor.email = email
        proveedor.ultimaEntrega = ultimaEntrega

        proveedor.save()

        return redirect('/Proveedores/')
    
def eliminarProveedor(request, codigo, url):
    proveedor = Proveedores.objects.get(codigoProveedor=int(codigo))
    proveedor.delete()

    return redirect(f'/{url}/')

#Exportar Excel Bodega

def excelProveedores(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Proveedores.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:G1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Proveedores'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Nombre', 'Telefono', 'Direccion', 'Email', 'Ultima Entrega']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoProveedor = registro.codigoProveedor
            nombreProveedor = registro.nombreProveedor
            telefono = registro.telefono
            direccion = registro.direccion
            email = registro.email
            ultimaEntrega = registro.ultimaEntrega

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoProveedor).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=nombreProveedor).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=telefono).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=direccion).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=email).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=ultimaEntrega).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 7):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoProveedores/')

def pdfProveedores(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Proveedores")

    pdf.tableProveedores()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoProveedores/')

#Categorias

#Home
def categorias(request, codigo=None):
    listarCategorias = Categorias.objects.all()
    
    return render(request, "Categorias/panel.html", {'Categorias':listarCategorias,'Codigo':codigo})

def listadoCategorias(request):
    listarCategorias = Categorias.objects.all()
    
    return render(request, "Categorias/listado.html", {'Categorias': listarCategorias})

def registrarCategoria(request):
    if request.method == 'POST':
        codigoCategoria = request.POST['primaryKey']
        nombreCategoria = request.POST['nombreCategoria']

        imagen = None
        if 'imageInput' in request.FILES:
            imagen = request.FILES['imageInput']

        categoria = Categorias.objects.create(
            codigoCategorias=codigoCategoria,
            nombreCategoria=nombreCategoria,
            imagenCategoria=imagen,
        )

        categoria.save()

        return redirect('/Categorias/')
    else:
        return redirect('/')

def obtener_categoria(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Productos que coincida con el código
        categoria = get_object_or_404(Categorias, codigoCategorias=codigo)
        data = {
            'nombreCategoria': categoria.nombreCategoria,
            'imagen': categoria.imagenCategoria.url if categoria.imagenCategoria else None,
        }
        return JsonResponse({'data': data})

def editarCategoria(request, codigo):
    categoria = Categorias.objects.get(codigoCategorias=int(codigo))

    if request.method == 'POST':

        nombreCategoria = request.POST['nombreCategoria']

        if 'imageInput' in request.FILES:
            imagen = request.FILES['imageInput']

            if categoria.imagenCategoria != "":
                ruta_foto = "Files/" + str(categoria.imagenCategoria)
                os.remove(ruta_foto)

            ruta_foto = "Files/Categorias/" + str(imagen)
            with open(ruta_foto, 'wb') as f:
                for chunk in imagen.chunks():
                    f.write(chunk)

            categoria.imagenCategoria = "/Categorias/" + str(imagen)

        categoria.nombreCategoria = nombreCategoria

        categoria.save()

        return redirect('/Categorias/')

def eliminarCategoria(request, codigo, url):

    #Elimina al trabajador

    registro = Categorias.objects.get(codigoCategorias=int(codigo))

    if registro.imagenCategoria != "":
        ruta_foto = "Files/" + str(registro.imagenCategoria)
        os.remove(ruta_foto)

    registro.delete()

    #Verificamos la url a la que redirige
    return redirect(f'/{url}/')

#Exportar Excel Categorias

def excelCategorias(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Categorias.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:B1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Categorias'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo Categoria', 'Nombre Categoria']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoCategoria = registro.codigoCategorias
            nombreCategoria = registro.nombreCategoria

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoCategoria).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=nombreCategoria).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 2):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoCategorias/')

def pdfCategorias(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Categorias")

    pdf.tableCategorias()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoCategorias/')