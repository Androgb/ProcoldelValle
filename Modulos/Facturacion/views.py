from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Pagos, ProductoCompra
from ..Inventario.models import *
from ..Usuarios.models import Clientes
from ..Tienda.models import *
from .models import *
from django.shortcuts import *
import decimal
import os
import random
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.conf import settings

#Importamos Librerias
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF, HTMLMixin




def compras(request, codigo=None):
    listarCompras = Compras.objects.all()
    listarTrabajadores = Trabajadores.objects.all()
    listarCajas = Cajas.objects.all()

    return render(request, "Compras/panel.html", {'Compras': listarCompras, 'Trabajadores': listarTrabajadores, 'Cajas': listarCajas, 'Codigo': codigo})

def listadoCompras(request):
    listarCompras = Compras.objects.all()
    listarTrabajadores = Trabajadores.objects.all()
    return render(request, "Compras/listado.html", {'Compras': listarCompras, 'Trabajadores': listarTrabajadores})

def registrarCompra(request):
    if request.method == 'POST':
        codigoCompra = request.POST['primaryKey']
        trabajador = request.POST['trabajador']
        servicio = request.POST['servicio']
        descripcion = request.POST['descripcion']
        gasto = request.POST['gasto']
        codigoCaja = request.POST['caja']
        caja = Cajas.objects.get(codigoCaja=int(codigoCaja))

        comprobante = None
        if 'imageInput' in request.FILES:
            comprobante = request.FILES['imageInput']

        tra = Trabajadores.objects.get(cedula=trabajador)

        compra = Compras.objects.create(
            codigoCompra=codigoCompra,
            trabajador=tra,
            servicio=servicio,
            descripcion=descripcion,
            Gasto=gasto,
            fecha=datetime.now().date(),
            caja = caja,
            comprobante= comprobante
        )

        #Crear el movimiento

        codigoMovimiento = random.randint(1, 3000)
        movimientoTipo = "Salida de Dinero"

        movimiento = Movimientos.objects.create(
            codigoMovimiento=codigoMovimiento,
            Caja=caja,
            Movimiento=movimientoTipo,
            Efectivo=gasto,
            Motivo=descripcion,
            fecha=datetime.now().date()
        )

        #Actualiza la caja el efectivo

        efectivoCaja = caja.Efectivo

        efectivoCajaActualizado = efectivoCaja - int(gasto)

        caja.Efectivo = efectivoCajaActualizado

        caja.save()

        return redirect('/Compras/')
    else:
        return redirect('/Compras')

def obtener_compra(request, codigo):
    if request.method == 'GET':
        compra = get_object_or_404(Compras, codigoCompra=codigo)
        data = {
            'trabajador': compra.trabajador.cedula,
            'servicio': compra.servicio,
            'descripcion': compra.descripcion,
            'gasto': str(compra.Gasto),
            'comprobante': compra.comprobante.url if compra.comprobante else None,
        }
        return JsonResponse({'data': data})

def editarCompra(request, codigo):
    compra = Compras.objects.get(codigoCompra=int(codigo))

    if request.method == 'POST':
        trabajador = request.POST['trabajador']
        servicio = request.POST['servicio']
        descripcion = request.POST['descripcion']
        gasto = request.POST['gasto']

        if 'imageInput' in request.FILES:
            comprobante = request.FILES['imageInput']

            if compra.comprobante != "":
                ruta_foto = "Files/" + str(compra.comprobante)
                os.remove(ruta_foto)

            ruta_foto = "Files/Productos/" + str(comprobante)
            with open(ruta_foto, 'wb') as f:
                for chunk in comprobante.chunks():
                    f.write(chunk)

            compra.comprobante = "/Productos/" + str(comprobante)

        trabajador = Trabajadores.objects.get(cedula=trabajador)

        compra.trabajador = trabajador
        compra.servicio = servicio
        compra.descripcion = descripcion
        compra.Gasto = gasto
        compra.fecha = datetime.now().date()

        compra.save()

        return redirect('/Compras/')

def eliminarCompra(request, codigo, url):
    compra = Compras.objects.get(codigoCompra=int(codigo))
    if compra.comprobante != "":
        ruta_foto = "Files/" + str(compra.comprobante)
        os.remove(ruta_foto)

    codigoCaja = compra.caja.codigoCaja

    caja = Cajas.objects.get(codigoCaja=int(codigoCaja))

    efectivoCaja = caja.Efectivo

    compra.delete()
    return redirect(f'/{url}/')

#Excel Compras

def excelCompras(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Compras.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:F1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Compras'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Trabajador', 'Servicio', 'Gasto', 'Descripcion', 'Fecha']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoCompra = registro.codigoCompra
            trabajador = registro.trabajador.nombreCompleto
            servicio = registro.servicio
            gasto = registro.Gasto
            descripcion = registro.descripcion
            fecha = str(registro.fecha)

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoCompra).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=trabajador).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=servicio).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=gasto).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=descripcion).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=fecha).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 6):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoCompras/')

def pdfCompras(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Compras")

    pdf.tableCompras()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoCompras/')


def movimientos(request, codigo=None):
    listarMovimientos = Movimientos.objects.all()
    listarCajas = Cajas.objects.all()

    return render(request, "Movimientos/panel.html", {'Movimientos': listarMovimientos, 'Cajas': listarCajas, 'Codigo': codigo})

def listadoMovimientos(request):
    listarMovimientos = Movimientos.objects.all()
    listarCajas = Cajas.objects.all()
    return render(request, "Movimientos/listado.html", {'Movimientos': listarMovimientos,'Cajas': listarCajas})

def registrarMovimientos(request):
    if request.method == 'POST':
        codigoMovimiento = request.POST['primaryKey']
        caja_id = request.POST['caja']
        movimientoTipo = request.POST['movimiento']
        efectivo = request.POST['efectivo']
        motivo = request.POST['motivo']


        caja = Cajas.objects.get(codigoCaja=caja_id)

        movimiento = Movimientos.objects.create(
            codigoMovimiento=codigoMovimiento,
            Caja=caja,
            Movimiento=movimientoTipo,
            Efectivo=efectivo,
            Motivo=motivo,
            fecha=datetime.now().date()
        )

        #Actualiza la caja el efectivo

        efectivoCaja = caja.Efectivo

        if movimientoTipo == "Entrada de Dinero":
            efectivoCajaActualizado = efectivoCaja + int(efectivo)
        else:
            efectivoCajaActualizado = efectivoCaja - int(efectivo)

        caja.Efectivo = efectivoCajaActualizado

        caja.save()

        movimiento.save()

        return redirect('/Movimientos/')
    else:
        return redirect('/Movimientos/')

def obtener_movimientos(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Movimientos que coincida con el código
        movimiento = get_object_or_404(Movimientos, codigoMovimiento=codigo)
        data = {
            'Caja': movimiento.Caja.codigoCaja,  # Actualiza esta línea
            'Movimiento': movimiento.Movimiento,
            'Efectivo': movimiento.Efectivo,
            'Motivo': movimiento.Motivo
        }
        return JsonResponse({'data': data})

def editarMovimiento(request, codigo):
    movimiento = Movimientos.objects.get(codigoMovimiento=int(codigo))

    if request.method == 'POST':
        caja_id = request.POST['caja']
        movimiento_desc = request.POST['movimiento']
        efectivo = request.POST['efectivo']
        motivo = request.POST['motivo']

        caja = Cajas.objects.get(codigoCaja=caja_id)

        movimiento.Caja = caja
        movimiento.Movimiento = movimiento_desc
        movimiento.Efectivo = efectivo
        movimiento.Motivo = motivo
        movimiento.fecha = datetime.now().date()

        movimiento.save()

        return redirect('/Movimientos/')

def eliminarMovimientos(request, codigo, url):
    # Elimina el movimiento
    movimiento = Movimientos.objects.get(codigoMovimiento=int(codigo))

    movimientoTipo = movimiento.Movimiento
    codigoCaja = movimiento.Caja.codigoCaja

    caja = Cajas.objects.get(codigoCaja=int(codigoCaja))

    efectivoCaja = caja.Efectivo

    if movimientoTipo == "Entrada de Dinero":
        efectivoCajaActualizado = efectivoCaja - movimiento.Efectivo
    else:
        efectivoCajaActualizado = efectivoCaja + movimiento.Efectivo

    caja.Efectivo = efectivoCajaActualizado

    caja.save()

    movimiento.delete()

    # Verificamos la URL a la que redirige
    return redirect(f'/{url}/')

def excelMovimientos(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Movimientos.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:E1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Movimientos'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Numero Caja', 'Movimiento', 'Efectivo', 'Motivo']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoMovimiento = registro.codigoMovimiento
            numeroCaja = registro.Caja.numeroCaja
            movimiento = registro.Movimiento
            efectivo = registro.Efectivo
            motivo = registro.Motivo

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoMovimiento).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=numeroCaja).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=movimiento).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=efectivo).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=motivo).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 6):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoMovimientos/')

def pdfMovimientos(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Movimientos")

    pdf.tableMovimientos()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoMovimientos/')


from django.db.models import Sum, Case, When, IntegerField

def cajas(request):
    listarCajas = Cajas.objects.all()
    listarVentas = Pagos.objects.all()
    listarTrabajadores = Trabajadores.objects.all()
    listarMovimientos = Movimientos.objects.all()

    today = Movimientos.objects.latest('fecha').fecha
    movimientos = Movimientos.objects.filter(fecha=today)

    sumaEfectivoEntrada = movimientos.filter(Movimiento='Entrada de Dinero').aggregate(total=Sum('Efectivo'))['total']
    sumaEfectivoSalida = movimientos.filter(Movimiento='Salida de Dinero').aggregate(total=Sum('Efectivo'))['total']

    # Si no hay movimientos de entrada o salida, establecer los valores en 0
    sumaEfectivoEntrada = sumaEfectivoEntrada or 0
    sumaEfectivoSalida = sumaEfectivoSalida or 0

    return render(request, "Cajas/panel.html", {'Cajas': listarCajas,
                                                'Ventas': listarVentas,
                                                'Trabajadores': listarTrabajadores,
                                                'Movimientos': listarMovimientos,
                                                'Entra': sumaEfectivoEntrada,
                                                'Sale': sumaEfectivoSalida
                                                })




def cajas1(request, codigo=None):
    
    listarCajas = Cajas.objects.all()
    
    return render(request, "Cajas/cajas.html", {'Cajas':listarCajas,'Codigo':codigo})


def listadoCajas(request):
    
    listarCajas = Cajas.objects.all()
    
    return render(request, "Cajas/listado.html", {'Cajas':listarCajas})

def registrarCaja(request):
    if request.method == 'POST':
        codigoCaja = request.POST['primaryKey']
        numeroCaja = request.POST['numeroCaja']
        estadoCaja = request.POST['estado']
        efectivoCaja = request.POST['efectivo']

        caja = Cajas.objects.create(
            codigoCaja=codigoCaja,
            numeroCaja=numeroCaja,
            Estado=estadoCaja,
            Efectivo=efectivoCaja,
        )

        caja.save()

        return redirect('/Cajas/')
    else:
        return redirect('/')

def obtener_caja(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Cajas que coincida con el código
        caja = get_object_or_404(Cajas, codigoCaja=codigo)
        data = {
            'numeroCaja': caja.numeroCaja,
            'Estado': caja.Estado,
            'Efectivo': caja.Efectivo,
        }
        return JsonResponse({'data': data})

def editarCaja(request, codigo):
    #caja = Cajas.objects.get(codigoCaja=int(codigo))

    if request.method == "POST":
        
        return redirect('/Cajas/')

def eliminarCaja(request, codigo, url):
    # Elimina la caja
    caja = Cajas.objects.get(codigoCaja=int(codigo))
    caja.delete()

    # Verificamos la URL a la que redirige
    return redirect(f'/{url}/')

def excelCajas(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Cajas.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:D1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Cajas'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Numero Caja', 'Estado', 'Efectivo']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoCaja = registro.codigoCaja
            numeroCaja = registro.numeroCaja
            efectivo = registro.Efectivo
            estado = registro.Estado

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoCaja).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=numeroCaja).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=efectivo).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=estado).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 4):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoCajas/')

def pdfCajas(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Cajas")

    pdf.tableCajas()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoCajas/')


def shop(request):
    listarCategorias = Categorias.objects.all()
    listarProductos = Productos.objects.all()
    cedula = request.session.get("Cedula")
    if cedula is None:
        listarCarrito = 0
        return redirect("/login/")
    else:
        listarCarrito = CarritoCompras.objects.filter(cliente=int(cedula))
    # Calcular la suma de los precios totales
        total_carrito = sum(item.precioTotal for item in listarCarrito)

    return render(request, "Eccomerce/shop.html", {'Categorias':listarCategorias, 'Productos':listarProductos, 'Carrito': listarCarrito, 'TotalCarrito': total_carrito})

def perfilUsuarios(request):
    cedula = request.session.get("Cedula")
    
    if cedula is None:
        # Redireccionar a la página deseada cuando no se encuentra el valor en la sesión
        listarCarrito = 0
        return redirect("/login/")
    else:
        listarCarrito = CarritoCompras.objects.filter(cliente=int(cedula))

        # Calcular la suma de los precios totales
        total_carrito = sum(item.precioTotal for item in listarCarrito)

    return render(request, "Eccomerce/perfil.html", {'Carrito': listarCarrito, 'TotalCarrito': total_carrito})

def actualizarFotoUsuarios(request):
    if request.method == 'POST':

        cedula = request.POST['cedulaPhoto']
        fotoPerfil = request.FILES['photo']

        usuario = Clientes.objects.get(cedula=int(cedula))  

        # Elimina la foto de perfil anterior
        if usuario.fotoPerfil != "":
            ruta_foto = "Files/" + str(usuario.fotoPerfil)
            os.remove(ruta_foto)

        # Guarda la nueva foto de perfil
        ruta_foto = "Files/Avatars/Usuarios/" + str(fotoPerfil)
        with open(ruta_foto, 'wb') as f:
            for chunk in fotoPerfil.chunks():
                f.write(chunk)

        # Actualiza la ruta de la foto de perfil en la base de datos
        usuario.fotoPerfil = "Avatars/Usuarios/" + str(fotoPerfil)

        usuario.save()

        #Actualizamos la variable de sesion   

        request.session['Foto'] = str(usuario.fotoPerfil)

        return redirect("/perfilUsuarios/")

#Actualizar Informacion (Perfil)
def actualizarInformacionUsuario(request):
    if request.method == 'POST':
        #Obtenemos los datos del formulario
        cedula = request.POST['cedula']
        nombreCompleto = request.POST['nombreCompleto']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']

        #Actualizamos el trabajador
        usuario = Clientes.objects.get(cedula=int(cedula))

        usuario.nombreCompleto = nombreCompleto
        usuario.telefono = telefono
        usuario.direccion = direccion
        usuario.email = email

        usuario.save()

        #Actualizamos las variables de sesion

        request.session['Nombre'] = usuario.nombreCompleto
        request.session['Telefono'] = usuario.telefono
        request.session['Direccion'] = usuario.direccion
        request.session['Email'] = usuario.email

        return redirect('/perfilUsuarios/')

def contacto(request):
    listarCarrito = CarritoCompras.objects.filter(cliente=int(request.session.get("Cedula")))

    # Calcular la suma de los precios totales
    total_carrito = sum(item.precioTotal for item in listarCarrito)

    return render(request, "Eccomerce/contact-shop.html", {'Carrito': listarCarrito, 'TotalCarrito': total_carrito})

def carrito(request):
    listarCarrito = CarritoCompras.objects.filter(cliente=int(request.session.get("Cedula")))

    # Calcular la suma de los precios totales
    total_carrito = sum(item.precioTotal for item in listarCarrito)
    total_carrito_iva = total_carrito + (total_carrito * decimal.Decimal('0.19')).quantize(decimal.Decimal('0.0'))    

    return render(request, "Eccomerce/shoping-cart.html", {'Carrito': listarCarrito, 'TotalCarrito': total_carrito, 'TotalCarritoIva': total_carrito_iva})

def productDetail(request, codigo):
    producto = get_object_or_404(Productos, codigoProducto=codigo)
    listarReseñas = Reseñas.objects.filter(producto=codigo)
    listarCarrito = CarritoCompras.objects.filter(cliente=int(request.session.get("Cedula")))

    # Calcular la suma de los precios totales
    total_carrito = sum(item.precioTotal for item in listarCarrito)

    return render(request, "Eccomerce/product-detail.html", {'Producto': producto, 'Reseñas': listarReseñas, 'Carrito': listarCarrito, 'TotalCarrito': total_carrito})

def ordenes(request):
    listarOrdenes = Ordenes.objects.all()
    return render(request, "Ordenes/panel.html", {'Ordenes':listarOrdenes})

def obtener_orden(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Ordenes que coincida con el código
        orden = get_object_or_404(Ordenes, codigoOrden=codigo)
        data = {
            'Estado': orden.estado,
        }
        return JsonResponse({'data': data})

def editarOrden(request, codigo):
    orden = Ordenes.objects.get(codigoOrden=int(codigo))

    
    if request.method == 'POST':
        estado = request.POST['estado']

        orden.estado = estado

        orden.save()

        return redirect('/Ordenes/')

#Exportar Excel Bodega

def excelOrdenes(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Ordenes.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:G1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Ordenes'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Cedula', 'Nombre', 'Fecha', 'Total', 'Metodo Pago', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoOrden = registro.codigoOrden
            cedula = registro.cliente.cedula
            nombre = registro.cliente.nombreCompleto
            fecha = str(registro.fechaCreacion)
            total = registro.total
            metodoPago = registro.metodoPago
            estado = registro.estado

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoOrden).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=cedula).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=nombre).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=fecha).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=total).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=metodoPago).font = Font(name='Arial')
            sheet.cell(row=row_num, column=7, value=estado).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 7):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/Ordenes/')

def pdfOrdenes(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Ordenes")

    pdf.tableOrdenes()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/Ordenes/')

class PDF(FPDF):
    # Método para crear el encabezado
    def encabezado(self, pagina):
        # Agregamos la imagen del logo izquierdo
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Astrocode.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 10, 8, 28)
        # Agregamos la imagen del logo derecho
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Logo Empresa.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 167, 8, 33)

        # Agregamos el título centrado y en negrita
        # Agregamos el título centrado y en negrita
        self.set_font('Arial', 'B', 24)
        self.set_text_color(88, 129, 87)
        self.cell(190, 15, 'Listado de ' + pagina, 0, 0, 'C')
        self.ln(20)

        # Agregamos el subtitulo centrado y en negrita
        # Obtener la fecha y hora actuales
        now = datetime.now()

        # Formatear la fecha y hora como una cadena con el formato deseado
        fecha_hora = now.strftime("%d/%m/%Y %H:%M:%S")
        
        # Agregar el subtitulo centrado y en negrita
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(0, -10, f"Exportado el: {fecha_hora}", 0, 0, 'C')
        self.ln(10)

        

    #Metodo para crear la tabla
    def tableOrdenes(self):
        #Registros Base de datos
        registros = Ordenes.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [10, 37, 64, 91, 118, 145, 172]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Cedula', 'Nombre', 'Fecha', 'Total', 'Metodo Pago', 'Estado']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoOrden = str(registro.codigoOrden)
            cedula = str(registro.cliente.cedula)
            nombre = registro.cliente.nombreCompleto
            fecha = str(registro.fechaCreacion)
            total = str(registro.total)
            metodoPago = registro.metodoPago
            estado = registro.estado

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoOrden, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, cedula, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, nombre, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, fecha, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, total, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.cell(27, 11, metodoPago, 1, 0, 'C', False, 0)

            self.set_x(x_coords[6])
            self.cell(27, 11, estado, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    #Metodo para crear la tabla
    def tableMovimientos(self):
        #Registros Base de datos
        registros = Movimientos.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [34, 61, 88, 115, 142]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Numero Caja', 'Movimiento', 'Efectivo', 'Motivo']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoMovimiento = str(registro.codigoMovimiento)
            caja = str(registro.Caja.numeroCaja)
            movimiento = registro.Movimiento
            efectivo = str(registro.Efectivo)
            motivo = str(registro.Motivo)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoMovimiento, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, caja, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.cell(27, 11, movimiento, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, efectivo, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.cell(27, 11, motivo, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableCompras(self):
        #Registros Base de datos
        registros = Compras.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [20, 47, 74, 101, 128, 155]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Trabajador', 'Servicio', 'Gasto', 'Descripcion', 'Fecha']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoCompra = str(registro.codigoCompra)
            trabajador = registro.trabajador.nombreCompleto
            servicio = registro.servicio
            gasto = str(registro.Gasto)
            descripcion = registro.descripcion
            fecha = str(registro.fecha)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoCompra, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, trabajador, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, servicio, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, gasto, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, descripcion, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, fecha, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableCajas(self):
        #Registros Base de datos
        registros = Cajas.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [43, 70, 97, 124]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Numero Caja', 'Estado', 'Efectivo']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoCaja = str(registro.codigoCaja)
            numeroCaja = str(registro.numeroCaja)
            estado = registro.Estado
            efectivo = str(registro.Efectivo)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoCaja, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, numeroCaja, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.cell(27, 11, estado, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, efectivo, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableVentasTotales(self):
        #Registros Base de datos
        registros = Pagos.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [25, 52, 79, 106, 133, 160]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo', 'Cliente', 'Caja', 'Metodo Pago', 'Fecha', 'Total']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            codigoFactura = str(registro.codigoFactura)
            cliente = registro.cliente.nombreCompleto
            caja = "Caja #" + str(registro.caja.numeroCaja)
            metodoPago = str(registro.metodoPago)
            fecha = str(registro.fechaPedido)
            total = str(registro.precioTotal)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 8)
            self.cell(27, 11, codigoFactura, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, cliente, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, caja, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, metodoPago, 1, 0, 'C', False, 0)

            self.set_x(x_coords[4])
            self.set_font('Arial', '', 5)
            self.cell(27, 11, fecha, 1, 0, 'C', False, 0)

            self.set_x(x_coords[5])
            self.set_font('Arial', '', 8)
            self.cell(27, 11, total, 1, 0, 'C', False, 0)

            self.ln()
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    def tableVentas(self, codigo):
        #Registros Base de datos
        registros = ProductoCompra.objects.filter(pago=int(codigo))

        #Subtotal

        subtotalCarrito = sum(item.subtotal for item in registros)
        iva = (subtotalCarrito * decimal.Decimal('0.19')).quantize(decimal.Decimal('0.0'))    
        total = subtotalCarrito + iva

        #Busca el pago que coincida con el codigo
        pago = Pagos.objects.get(codigoFactura=int(codigo))

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [55, 82, 109, 136]
        x_start = x_coords[0]  # Guardar la posición x inicial

        # Calcular el ancho total de la tabla
        table_width = x_coords[-1] - x_start + 27

        # Centrar los encabezados
        self.set_x(x_start + (table_width - self.get_string_width('Codigo')) / 2)
        for i, header in enumerate(['Codigo Factura', 'Producto', 'Cantidad', 'Subtotal']):
            if i == 0:
                self.set_xy(x_start, self.get_y())  # Usar la posición x inicial
            else:
                self.set_xy(x_coords[i], self.get_y())

            self.cell(27, 11, header, 1, 0, 'C', True, 0)

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            producto = registro.producto.nombreProducto
            cantidad = str(registro.cantidad)
            subtotal = str(registro.subtotal)

            # Imprimir cada celda en la posición adecuada
            
            # Establecer la posición x al inicio de cada fila
            self.set_x(x_start)

            self.set_font('Arial', '', 9)
            self.cell(27, 11, codigo, 1, 0, 'C', False, 0)

            self.set_x(x_coords[1])
            self.cell(27, 11, producto, 1, 0, 'C', False, 0)

            self.set_x(x_coords[2])
            self.cell(27, 11, cantidad, 1, 0, 'C', False, 0)

            self.set_x(x_coords[3])
            self.cell(27, 11, subtotal, 1, 0, 'C', False, 0)

            self.ln()

        self.ln(10)  # Salto de línea para dejar espacio

        # Establecer la coordenada x inicial del contenedor
        x_start_contenedor = 10

        # Establecer la coordenada y inicial del contenedor
        y_start_contenedor = self.get_y()

        # Establecer el ancho total del contenedor
        contenedor_width = 192 - x_start_contenedor

        # Establecer la altura del contenedor
        contenedor_height = 10

        # Establecer el color de texto para el contenedor
        self.set_text_color(0, 0, 0)  # Color de texto negro

        # Establecer la fuente y el estilo para el título del contenedor
        self.set_font('Arial', 'B', 14)

        # Imprimir el título de la columna izquierda del contenedor sin color de fondo
        self.set_xy(x_start_contenedor, y_start_contenedor)
        self.cell(contenedor_width / 2, contenedor_height, 'Informacion Cliente', 0, 0, 'L')

        # Imprimir el título de la columna derecha del contenedor sin color de fondo
        self.set_xy(x_start_contenedor + contenedor_width / 2, y_start_contenedor)
        self.cell(contenedor_width / 2, contenedor_height, 'Total Factura', 0, 0, 'R')

        # Establecer la fuente y el estilo para los valores del contenedor
        self.set_font('Arial', '', 12)

        # Imprimir el valor de la columna izquierda del contenedor alineado a la izquierda
        self.set_xy(x_start_contenedor, y_start_contenedor + contenedor_height)
        self.cell(contenedor_width / 2, contenedor_height, f'Cedula: {pago.cliente.cedula}', 0, 0, 'L')

        # Imprimir el valor de la columna derecha del contenedor alineado a la derecha
        self.set_xy(x_start_contenedor + contenedor_width / 2, y_start_contenedor + contenedor_height)
        self.cell(contenedor_width / 2, contenedor_height, f'Subtotal: + ${subtotal} COP', 0, 0, 'R')

        # Agregar tres filas adicionales al contenedor
        filas_adicionales = [
            (f'Nombre: {pago.cliente.nombreCompleto}', f'Iva: + ${iva} COP'),
            (f'Caja: # {pago.caja.numeroCaja}', f'Total: ${total} COP'),
            (f'Metodo de Pago: {pago.metodoPago}', '')
        ]

        for fila in filas_adicionales:
            self.ln()  # Salto de línea
            self.set_x(x_start_contenedor)
            self.cell(contenedor_width / 2, contenedor_height, fila[0], 0, 0, 'L')  # Columna izquierda
            self.set_x(x_start_contenedor + contenedor_width / 2)
            self.cell(contenedor_width / 2, contenedor_height, fila[1], 0, 0, 'R')  # Columna derecha vacía si es necesario
        
        if self.get_y() >= self.page_break_trigger:
            self.add_page()
            self.set_xy(x_start, y_start)  # Restaurar la posición inicial para la siguiente página

    # Método para crear el footer
    def footer(self):
        # Agregamos el texto del footer en gris
        self.set_y(-15)
        self.set_font('Arial', '', 8)
        self.set_text_color(127, 127, 127)
        self.cell(0, 10, 'Procol del Valle © Astrocode | Página ' + str(self.page_no()), 0, 0, 'C')


def ventas(request, codigoVenta=None):
    listarProductos = Productos.objects.all()
    listarClientes = Clientes.objects.all()
    listarCajas = Cajas.objects.all()

    if codigoVenta is None:
        return render(request, "Ventas/venta.html", {'Productos': listarProductos, 'Clientes': listarClientes, 'Codigo':codigoVenta,'Cajas':listarCajas})
    else:
        listarProductosCompra = ProductoCompra.objects.filter(pago=int(codigoVenta))
        listarVentas = Pagos.objects.get(codigoFactura = int(codigoVenta))

        #Subtotal

        subtotalCarrito = sum(item.subtotal for item in listarProductosCompra)
        iva = (subtotalCarrito * decimal.Decimal('0.19')).quantize(decimal.Decimal('0.0'))    
        total = subtotalCarrito + iva

        return render(request, "Ventas/venta.html", {'Ventas':listarVentas, 'Productos': listarProductos, 'ProductosCompra': listarProductosCompra, 'Subtotal': subtotalCarrito, 'Iva': iva, 'Total': total, 'Clientes': listarClientes, 'Codigo':codigoVenta,'Cajas':listarCajas})

def actualizarVenta(request, codigoFactura):
    
    factura = Pagos.objects.get(codigoFactura=int(codigoFactura))

    #Obtenemos los datos
    cajaCodigo = request.POST['caja']
    caja = Cajas.objects.get(codigoCaja=int(cajaCodigo))
    cedula = request.POST['cliente']
    usuario = Clientes.objects.get(cedula=int(cedula))
    metodo_pago = request.POST['metodo_pago']

    listarProductosCompra = ProductoCompra.objects.filter(pago=int(codigoFactura))

    #Subtotal

    subtotalCarrito = sum(item.subtotal for item in listarProductosCompra)
    iva = (subtotalCarrito * decimal.Decimal('0.19')).quantize(decimal.Decimal('0.0'))    
    total = subtotalCarrito + iva

    if factura.precioTotal is None:
        total_entero = 0
    else:
        total_entero = int(factura.precioTotal)

    #Actualizamos los datos
    factura.cliente = usuario
    factura.caja = caja
    factura.metodoPago = metodo_pago
    factura.precioTotal = total

    factura.save()

    #Valida si existe el movimiento o no

    movimientoSearch = Movimientos.objects.filter(Movimiento="Entrada de Dinero", Efectivo=total_entero, fecha=datetime.now().date())

    if movimientoSearch.exists():
        movimiento = movimientoSearch.first()

        #Actualiza
        movimiento.Caja = caja
        movimiento.Efectivo = total

        #Actualiza la caja el efectivo

        efectivoCaja = caja.Efectivo

        if total > total_entero:
            efectivoCajaActualizado = efectivoCaja + int(total - total_entero)
        else:
            efectivoCajaActualizado = efectivoCaja - int(total_entero - total)

        caja.Efectivo = efectivoCajaActualizado

        caja.save()

        movimiento.save()

    else:

        #Crear el movimiento

        codigoMovimiento = random.randint(1, 3000)
        movimientoTipo = "Entrada de Dinero"

        movimiento = Movimientos.objects.create(
            codigoMovimiento=codigoMovimiento,
            Caja=caja,
            Movimiento=movimientoTipo,
            Efectivo=total,
            Motivo="Venta de Producto",
            fecha=datetime.now().date()
        )

        #Actualiza la caja el efectivo

        efectivoCaja = caja.Efectivo

        efectivoCajaActualizado = efectivoCaja + int(total)

        caja.Efectivo = efectivoCajaActualizado

        caja.save()
 
    return redirect("/Ventas/" + str(codigoFactura))

def agregarProductoCompra(request, codigo, codigoFactura):
    
    if request.method == "POST":
        

        #Valido si la factura existe o no
        if int(codigoFactura) == 0:
            #Crea la factura y la deja vacia

            fechaCreacion = datetime.now()

            # Formato personalizado: "YYYY-MM-DD HH:MM:SS"
            fechaFormateada = fechaCreacion.strftime("%Y-%m-%d %H:%M:%S")

            factura = Pagos.objects.create(
                codigoFactura = None,
                cliente = None,
                fechaPedido = fechaFormateada, 
                precioTotal = None,
                metodoPago = None,
                caja = None,
            )

            url = "/Ventas/" + str(factura.codigoFactura)

            codigoFacturaProducto = factura.codigoFactura;
        
        else:
            url = "/Ventas/" + str(codigoFactura)

            codigoFacturaProducto = codigoFactura;

        #Agregamos el producto a producto_compra

        producto = Productos.objects.get(codigoProducto=int(codigo))
        pago = Pagos.objects.get(codigoFactura=int(codigoFacturaProducto))

        productoCompra = ProductoCompra.objects.create(
                codigoProductoCompra = None,
                producto = producto,
                pago = pago, 
                cantidad = 1,
                subtotal = producto.precioUnidad,
        )
        
        #Actualiza el stock del producto

        stock = producto.stock

        stockActualizado = stock - 1

        producto.stock = stockActualizado

        producto.save()

        return redirect(url)

def actualizarProductoCompra(request, codigoProductoCompra):
    
    productoCompra = ProductoCompra.objects.get(codigoProductoCompra=int(codigoProductoCompra))

    #Obtenemos los datos
    cantidad = int(request.POST['cantidad'])
    precioUnidad = productoCompra.producto.precioUnidad;

    producto = Productos.objects.get(codigoProducto=int(productoCompra.producto.codigoProducto))

    stock = producto.stock
    
    if cantidad > int(productoCompra.cantidad):
        stockActualizado = stock - (cantidad - int(productoCompra.cantidad))
    else:
        stockActualizado = stock + (int(productoCompra.cantidad) - cantidad)
    
    producto.stock = stockActualizado

    producto.save()
    
    subtotal = cantidad * precioUnidad

    #Actualizamos los datos
    productoCompra.cantidad = cantidad
    productoCompra.subtotal = subtotal

    productoCompra.save() 

    return redirect("/Ventas/" + str(productoCompra.pago.codigoFactura))

def eliminarProductoCompra(request, codigoProductoCompra):
    productoCompra = ProductoCompra.objects.get(codigoProductoCompra=int(codigoProductoCompra))

    producto = Productos.objects.get(codigoProducto=int(productoCompra.producto.codigoProducto))

    stock = producto.stock

    stockActualizado = int(stock) + int(productoCompra.cantidad)

    producto.stock = stockActualizado

    producto.save()

    # Elimina el producto
    url = "/Ventas/" + str(productoCompra.pago.codigoFactura)

    productoCompra.delete()

    # Verificamos la URL a la que redirige
    return redirect(url)
        

def listadoVentas(request):
    listarVentas = Pagos.objects.all()
    
    return render(request, "Ventas/facturas.html", {'Ventas':listarVentas})

def obtener_venta(request, codigo):
    if request.method == 'GET':
        # Obtiene el elemento de la tabla Ordenes que coincida con el código
        pago = get_object_or_404(Pagos, codigoFactura=codigo)
        data = {}
        if pago.caja != None and pago.cliente != None and pago.metodoPago != None:
            data = {
                'Caja': pago.caja.codigoCaja,
                'Cliente': pago.cliente.cedula,
                'MetodoPago': pago.metodoPago,
            }
        
        return JsonResponse({'data': data})

def eliminarVenta(request, codigo, url):
    venta = Pagos.objects.get(codigoFactura=int(codigo))

    codigoCaja = venta.caja.codigoCaja

    caja = Cajas.objects.get(codigoCaja=int(codigoCaja))

    #Actualizamos caja

    efectivoCaja = caja.Efectivo

    efectivoCajaActualizado = efectivoCaja - int(venta.precioTotal)

    caja.Efectivo = efectivoCajaActualizado

    caja.save()

    #Eliminamos el movimiento

    movimientoSearch = Movimientos.objects.filter(Movimiento="Entrada de Dinero", Efectivo=int(venta.precioTotal), fecha=datetime.now().date())

    if movimientoSearch.exists():
        movimiento = movimientoSearch.first()

        movimiento.delete()

    venta.delete()

    # Verificamos la URL a la que redirige
    return redirect(f'/{url}/')

def pdfVentasIndividual(request, codigo):
    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Ventas")

    pdf.tableVentas(codigo)

    # Obtener la ruta de la carpeta "Descargas"
    descargas_folder = os.path.join(os.path.expanduser('~'), 'Downloads')

    # Guardar el archivo PDF en la carpeta "Descargas"
    pdf_output_path = os.path.join(descargas_folder, 'Factura Individual.pdf')
    pdf.output(pdf_output_path, 'F')

    return redirect('/Ventas/' + codigo)

def pdfVentas(request, fileName, directorySelected):
    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.encabezado("Ventas")

    pdf.tableVentasTotales()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoVentas/')

#Exportar Excel Bodega

def excelVentas(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Pagos.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:F1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Ventas'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Codigo', 'Cliente', 'Caja', 'Metodo Pago', 'Fecha', 'Total']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            codigoFactura = registro.codigoFactura
            cliente = registro.cliente.nombreCompleto
            caja = "Caja #" + str(registro.caja.numeroCaja)
            metodoPago = registro.metodoPago
            fecha = str(registro.fechaPedido)
            total = registro.precioTotal

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=codigoFactura).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=cliente).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=caja).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=metodoPago).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=fecha).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=total).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 6):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoVentas/')

def verifyDirectories(directory, file):
    #Crea la carpeta donde se almacenan los datos
    directory_path = os.path.join(os.path.join(os.environ['USERPROFILE']), directory)
    procol_path = os.path.join(directory_path, 'Procol del Valle')

    if not os.path.exists(procol_path):
        os.makedirs(procol_path)
        excels_path = os.path.join(procol_path, 'Excels')
        pdf_path = os.path.join(procol_path, 'PDF')
        os.makedirs(excels_path)
        os.makedirs(pdf_path)

    if file == "Excel":
        excels_path = os.path.join(procol_path, 'Excels')

        if not os.path.exists(excels_path):
            os.makedirs(excels_path)

        return excels_path
    elif file == "Pdf":
        pdfs_path = os.path.join(procol_path, 'PDF')

        if not os.path.exists(pdfs_path):
            os.makedirs(pdfs_path)
        
        return pdfs_path