from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.http import JsonResponse
import os
import datetime
from django.conf import settings

#Importamos Librerias
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from fpdf import FPDF, HTMLMixin

#Enviar Email
import uuid

import yagmail

emailFromAccount='procoldelvalleweb@gmail.com'
passwordFromAccount = 'wuubqrfjkefnycyp'

yag = yagmail.SMTP(user=emailFromAccount, password=passwordFromAccount)

#USUARIOS

def registrarUsuario(request):
    if request.method == 'POST':
        cedula = request.POST['cedula-register']
        nombreCompleto = request.POST['nombre-register']
        direccion = request.POST['direccion-register']
        email = request.POST['email-register']
        telefono = request.POST['telefono-register']
        contraseña = request.POST['password-register']

        #Verifica si se pasa una foto de perfil

        fotoPerfil = None
        
        if 'imageInput' in request.FILES:
            fotoPerfil = request.FILES['imageInput']

        usuario = Clientes.objects.create(
            cedula = cedula,
            nombreCompleto = nombreCompleto,
            direccion = direccion,
            email = email,
            telefono = telefono,
            contraseña = contraseña,
            fotoPerfil = fotoPerfil,
        )

        return redirect('/login/')    
        
def login(request):
    if request.method == 'POST':

        cedula = request.POST['cedula-login']
        contraseña = request.POST['password-login']

        #Verificamos que el usuario exista
        usuarios = Clientes.objects.filter(cedula=int(cedula), contraseña=contraseña)

        if usuarios.exists():
            usuario = usuarios.first()
            request.session['Cedula'] = usuario.cedula
            request.session['Nombre'] = usuario.nombreCompleto
            request.session['Foto'] = str(usuario.fotoPerfil)
            request.session['Telefono'] = usuario.telefono
            request.session['Direccion'] = usuario.direccion
            request.session['Email'] = usuario.email
            return redirect(settings.LOGIN_REDIRECT_URL_USER)    
        else:
            return redirect('/login/?message=La%20cedula/contraseña%20son%20incorrectos')
        
def loginTrabajador(request):
    if request.method == 'POST':

        cedula = request.POST['cedula-login']
        contraseña = request.POST['password-login']

        #Verificamos que el usuario exista
        trabajadores = Trabajadores.objects.filter(cedula=int(cedula), contraseña=contraseña)

        if trabajadores.exists():
            trabajador = trabajadores.first()
            request.session['Cedula'] = trabajador.cedula
            request.session['Nombre'] = trabajador.nombreCompleto
            request.session['Cargo'] = trabajador.cargo
            request.session['Telefono'] = trabajador.telefono
            request.session['Direccion'] = trabajador.direccion
            request.session['Email'] = trabajador.email
            request.session['Foto'] = str(trabajador.fotoPerfil)
            return redirect(settings.LOGIN_REDIRECT_URL)    
        else:
            return redirect('/login/?message=La%20cedula/contraseña%20son%20incorrectos')

def cerrarSesion(request):
    try:
        del request.session['Cedula']
        del request.session['Nombre']
        del request.session['Foto']

        if request.session.get('Cargo') != None:
            del request.session['Cargo']
        if request.session.get('Telefono') != None:
            del request.session['Telefono']
        if request.session.get('Direccion') != None:
            del request.session['Direccion']
        if request.session.get('Email') != None:
            del request.session['Email']
    except:
        return redirect(settings.LOGOUT_REDIRECT_URL)
    return redirect(settings.LOGOUT_REDIRECT_URL)

def sendEmail(request):
    if request.method == 'POST':
        email = request.POST['emailVerification']
        cedula = request.POST['cedulaVerification']

        #Verificamos que el usuario exista
        usuarios = Clientes.objects.filter(cedula=int(cedula))

        if usuarios.exists():
            usuario = usuarios.first()

            if usuario.email == email:
                # Generar un enlace único
                token = uuid.uuid4().hex

                usuario.token = token
                usuario.tokenUsed = False

                usuario.save()
                
                linkUnico = 'http://127.0.0.1:8000/resetPassword/' + cedula + '/' + usuario.token

                destinatarioEmail = [usuario.email]
                asuntoEmail = 'Restablecimiento de Contraseña - Procol del Valle'

                # Crear el contenido del correo
                banner_html = f'<img src="https://i.imgur.com/JfSZX40.png" style="width: 100%;" alt="Banner">'
                titulo_html = '<h1 style="font-family: Helvetica, Arial, sans-serif; font-size: 28px; font-weight: 400; margin-bottom: 0px;">¡Correo de Restablecimiento de Contraseña!</h1>'
                parrafo_html = '<p style="font-family: Helvetica, Arial, sans-serif; font-size: 16px; margin-top: 0px; margin-bottom: 0px;">Acabas de enviar una peticion para restablecer tu contraseña en el sistema de Procol del Valle. Si no fuiste tú quien realizó esta acción en nuestro software, es posible que otra persona esté usando tu cuenta. Revisa y protege tu cuenta ahora. Para restaurar tu contraseña ingresa al siguiente link unico:</p>'
                boton_html = f'''
                <a href="{linkUnico}" style="background-color: #189e76; color: white; border-radius: 20px; padding: 10px 20px; border: none; display: inline-flex; padding-left: 10px; padding-right: 0px; font-family: Helvetica, Arial, sans-serif; font-size: 16px; width: 180px; height: 30px; text-align: center; line-height: 10px; cursor: pointer; text-decoration: none;">
                    Restablecer Contraseña
                </a>
                '''
                contenido_html = f'{banner_html}\n{titulo_html}\n{parrafo_html}\n{boton_html}'    
                                    

                yag.send(to=destinatarioEmail, subject=asuntoEmail, contents=contenido_html)

                message = "El enlace para restablecer tu contraseña en nuestro software se ha enviado a tu correo de registro"

                # Reemplazar espacios en blanco con %20
                message_codified = message.replace(" ", "%20")

                return redirect(f'/login/?message={message_codified}&type=success')
            
            else:
                message = "El email no coincide con el email registrado en Procol del Valle"

                # Reemplazar espacios en blanco con %20
                message_codified = message.replace(" ", "%20")

                return redirect(f"/login/?message={message_codified}")

        else:
            message = "El usuario no fue encontrado, la cedula es incorrecta"

            # Reemplazar espacios en blanco con %20
            message_codified = message.replace(" ", "%20")

            return redirect(f"/login/?message={message_codified}")
        
def resetPasswordComplete(request):
    if request.method == 'POST':
        # Obtenemos los valores obtenidos por el formulario
        tokenReset = request.POST['password-token']
        cedulaReset = request.POST['password-cedula']
        passwordReset = request.POST['password-reset']

        #Busca el usuario en la BD que coincida con el token
        usuarios = Clientes.objects.filter(cedula=int(cedulaReset))

        if usuarios.exists():
            usuario = usuarios.first()

            if usuario.token == tokenReset and usuario.tokenUsed == 0:
                #Restaura la contraseña de la base de datos
                usuario.contraseña = passwordReset
                usuario.tokenUsed = True

                usuario.save()

                message = "Tu contraseña ha sido restablecida en nuestro software correctamente"

                # Reemplazar espacios en blanco con %20
                message_codified = message.replace(" ", "%20")

                return redirect(f'/login/?message={message_codified}&type=success')
            else:
                message = "El token ingresado no fue localizado en la base de datos o ya fue usado previamente, envia nuevamente una peticion para restaurar la contraseña"

                # Reemplazar espacios en blanco con %20
                message_codified = message.replace(" ", "%20")

                return redirect(f"/login/?message={message_codified}")
        else:
            message = "El usuario no fue encontrado, la cedula es incorrecta"

            # Reemplazar espacios en blanco con %20
            message_codified = message.replace(" ", "%20")

            return redirect(f"/login/?message={message_codified}")

#TRABAJADORES
#Home
def trabajadores(request, cedula=None):
    listar = Trabajadores.objects.all()
    return render(request, "trabajadores/panel.html", {'Trabajadores':listar, 'Cedula':cedula})

#Actualizar Informacion (Perfil)
def actualizarInformacion(request):
    if request.method == 'POST':
        #Obtenemos los datos del formulario
        cedula = request.POST['cedula']
        nombreCompleto = request.POST['nombreCompleto']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']

        #Actualizamos el trabajador
        trabajador = Trabajadores.objects.get(cedula=int(cedula))

        trabajador.nombreCompleto = nombreCompleto
        trabajador.telefono = telefono
        trabajador.direccion = direccion
        trabajador.email = email

        trabajador.save()

        #Actualizamos las variables de sesion

        request.session['Nombre'] = trabajador.nombreCompleto
        request.session['Telefono'] = trabajador.telefono
        request.session['Direccion'] = trabajador.direccion
        request.session['Email'] = trabajador.email

        return redirect('perfil')

def resetPasswordTrabajador(request):
    if request.method == 'POST':
        #Obtenemos los datos del formulario
        cedula = request.POST['cedulaHidden']
        password = request.POST['passwordField']

        #Actualizamos el trabajador
        trabajador = Trabajadores.objects.get(cedula=int(cedula))

        trabajador.contraseña = password

        trabajador.save()

        return redirect('perfil')

def actualizarFoto(request):
    if request.method == 'POST':

        cedula = request.POST['cedulaPhoto']
        fotoPerfil = request.FILES['photo']

        trabajador = Trabajadores.objects.get(cedula=int(cedula))  

        # Elimina la foto de perfil anterior
        if trabajador.fotoPerfil != "":
            ruta_foto = "Files/" + str(trabajador.fotoPerfil)
            os.remove(ruta_foto)

        # Guarda la nueva foto de perfil
        ruta_foto = "Files/Avatars/Trabajadores/" + str(fotoPerfil)
        with open(ruta_foto, 'wb') as f:
            for chunk in fotoPerfil.chunks():
                f.write(chunk)

        # Actualiza la ruta de la foto de perfil en la base de datos
        trabajador.fotoPerfil = "Avatars/Trabajadores/" + str(fotoPerfil)

        trabajador.save()

        #Actualizamos la variable de sesion   

        request.session['Foto'] = str(trabajador.fotoPerfil)

        return redirect("perfil")

        

#Listado trabajadores

def listadoTrabajadores(request):
    listar = Trabajadores.objects.all()
    return render(request, "trabajadores/listadoTrabajadores.html", {'Trabajadores': listar})

def registrarTrabajador(request):
    if request.method == 'POST':
        cedula = request.POST['primaryKey']
        nombreCompleto = request.POST['nombreCompleto']
        cargo = request.POST['cargo']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']
        contraseña = request.POST['contraseña']

        #Verifica si se pasa una foto de perfil

        fotoPerfil = None
        
        if 'imageInput' in request.FILES:
            fotoPerfil = request.FILES['imageInput']

        trabajador = Trabajadores.objects.create(
            cedula=cedula,
            nombreCompleto=nombreCompleto,
            cargo=cargo,
            telefono=telefono,
            direccion=direccion,
            email=email,
            contraseña=contraseña,
            fotoPerfil=fotoPerfil
        )        
        
        return redirect('/trabajadores/')    
    else:
        return redirect('/')
    
def obtener_trabajador(request, cedula):
    if request.method == 'GET':
        #Obtiene el elemento de la tabla Trabajadores que coincida con la cedula
        trabajador = get_object_or_404(Trabajadores, cedula=cedula)
        data = {
            'nombreCompleto': trabajador.nombreCompleto,
            'cargo': trabajador.cargo,
            'telefono': trabajador.telefono,
            'direccion': trabajador.direccion,
            'email': trabajador.email,
            'contraseña': trabajador.contraseña,
            'fotoPerfil': trabajador.fotoPerfil.url if trabajador.fotoPerfil else None,
        }
        return JsonResponse({'data': data})


def editarTrabajador(request, cedula):

    trabajador = Trabajadores.objects.get(cedula=int(cedula))

    if request.method == 'POST':

        #Obtener los datos del formulario

        nombreCompleto = request.POST['nombreCompleto']
        cargo = request.POST['cargo']
        telefono = request.POST['telefono']
        direccion = request.POST['direccion']
        email = request.POST['email']
        contraseña = request.POST['contraseña']

        if 'imageInput' in request.FILES:
            fotoPerfil = request.FILES['imageInput']

            # Elimina la foto de perfil anterior
            if trabajador.fotoPerfil != "":
                ruta_foto = "Files/" + str(trabajador.fotoPerfil)
                os.remove(ruta_foto)


        # Guarda la nueva foto de perfil
            ruta_foto = "Files/Avatars/Trabajadores/" + str(fotoPerfil)
            with open(ruta_foto, 'wb') as f:
                for chunk in fotoPerfil.chunks():
                    f.write(chunk)

            # Actualiza la ruta de la foto de perfil en la base de datos
            trabajador.fotoPerfil = "Avatars/Trabajadores/" + str(fotoPerfil)   

        trabajador.nombreCompleto = nombreCompleto
        trabajador.cargo = cargo
        trabajador.telefono = telefono
        trabajador.direccion = direccion
        trabajador.email = email
        trabajador.contraseña = contraseña

        trabajador.save()

        return redirect('/trabajadores/')
        

def eliminarTrabajador(request, cedula, url):

    #Elimina al trabajador

    trabajador = Trabajadores.objects.get(cedula=int(cedula))

    #Elimina la foto de perfil
    if trabajador.fotoPerfil != "":
        ruta_foto = "Files/" + str(trabajador.fotoPerfil)
        os.remove(ruta_foto)

    trabajador.delete()

    #Verificamos la url a la que redirige
    return redirect(f'/{url}/')

#Exportar Excel

def excelTrabajadores(request, fileName, directorySelected): 
    if request.method == 'GET':

        rutaExcel = verifyDirectories(directorySelected, "Excel")

        #Creamos el archivo excel con la libreria openpyxl

        #Obtenemos los registros de la base de datos
        registros = Trabajadores.objects.all()

        # crear un nuevo libro de trabajo y seleccionar la hoja activa
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambiar el nombre de la hoja a "Listado"
        sheet.title = 'Listado'

        # Fusionar celdas para el título
        sheet.merge_cells('A1:F1')

        # agregar un encabezado centrado y en negrita
        title = sheet['A1']
        title.value = 'Listado de Trabajadores'
        title.font = Font(size=20, bold=True, color='008000', name='Arial')
        title.alignment = Alignment(horizontal='center')

        # Establecer los estilos para el encabezado de la tabla
        header_fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
        header_font = Font(bold=True, color='FFFFFF', name="Arial", size=12)
        header_alignment = Alignment(horizontal='center')

        # agregar encabezados centrados y en negrita
        headers = ['Cedula', 'Nombre Completo', 'Cargo', 'Telefono', 'Direccion', 'Email']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Definir el ancho deseado para todas las celdas
        ancho_celda = 40

        # Recorrer todas las filas y celdas
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si la celda tiene información
                if cell.value:
                    # Establecer el mismo ancho para todas las celdas
                    sheet.column_dimensions[cell.column_letter].width = ancho_celda

        # Añadir los registros a partir de la fila 3
        for row_num, registro in enumerate(registros, start=3):
            # Obtener los valores de cada campo del registro
            cedula = registro.cedula
            nombre = registro.nombreCompleto
            cargo = registro.cargo
            telefono = registro.telefono
            direccion = registro.direccion
            email = registro.email

            # Añadir los valores de cada campo a las celdas correspondientes
            sheet.cell(row=row_num, column=1, value=cedula).font = Font(name='Arial')
            sheet.cell(row=row_num, column=2, value=nombre).font = Font(name='Arial')
            sheet.cell(row=row_num, column=3, value=cargo).font = Font(name='Arial')
            sheet.cell(row=row_num, column=4, value=telefono).font = Font(name='Arial')
            sheet.cell(row=row_num, column=5, value=direccion).font = Font(name='Arial')
            sheet.cell(row=row_num, column=6, value=email).font = Font(name='Arial')

            # Establecer el alineado central
            for col_num in range(1, 7):
                sheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

        # guardar el archivo Excel
        workbook.save(os.path.join(rutaExcel, fileName + '.xlsx'))
    
        return redirect('/listadoTrabajadores/')

def pdfTrabajadores(request, fileName, directorySelected):

    #Verificamos el directorio
    rutaPdf = verifyDirectories(directorySelected, "Pdf")

    #Creamos el PDF con Fpdf
    pdf = PDF()

    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    pdf.table()

    # Guardamos el PDF en la ruta del proyecto
    pdf.output(os.path.join(rutaPdf, fileName + '.pdf'), "F")

    return redirect('/listadoTrabajadores/')

class PDF(FPDF):
    # Método para crear el encabezado
    def header(self):
        # Agregamos la imagen del logo izquierdo
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Astrocode.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 10, 8, 28)
        # Agregamos la imagen del logo derecho
        ruta_imagen = os.path.join(settings.BASE_DIR, 'procoldelvalle', 'static', 'assets', 'images', 'Logo Empresa.png')
        ruta_imagen = ruta_imagen.replace('\\', '/')
        self.image(ruta_imagen, 167, 8, 33)

        # Agregamos el título centrado y en negrita
        self.set_font('Arial', 'B', 24)
        self.set_text_color(88, 129, 87)
        self.cell(0, 15, 'Listado de Trabajadores', 0, 0, 'C')
        self.ln(20)

        # Agregamos el subtitulo centrado y en negrita
        # Obtener la fecha y hora actuales
        now = datetime.datetime.now()

        # Formatear la fecha y hora como una cadena con el formato deseado
        fecha_hora = now.strftime("%d/%m/%Y %H:%M:%S")
        
        # Agregar el subtitulo centrado y en negrita
        self.set_font('Arial', 'B', 10)
        self.set_text_color(0, 0, 0)
        self.cell(0, -10, f"Exportado el: {fecha_hora}", 0, 0, 'C')
        self.ln(10)

    #Metodo para crear la tabla
    def table(self):
        #Registros Base de datos
        registros = Trabajadores.objects.all()

        # Agregar los encabezados de la tabla
        self.set_fill_color(88, 129, 87)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 10)

        x_coords = [42, 74, 106, 138, 170, 202]

        # Imprimir los encabezados de la tabla en la posición adecuada
        for i, header in enumerate(['Cedula', 'Nombre', 'Cargo', 'Telefono', 'Direccion', 'Email']):
            self.cell(32, 11, header, 1, 0, 'C', True, 0)
            self.set_xy(x_coords[i], self.get_y())

        self.ln()

        # Recorremos los registros y los agregamos a la tabla
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_xy(10, self.get_y())

        for registro in registros:
            cedula = str(registro.cedula)
            nombre = registro.nombreCompleto
            cargo = registro.cargo
            telefono = str(registro.telefono)
            direccion = registro.direccion
            email = registro.email

            # Imprimir cada celda en la posición adecuada
            for i, cell in enumerate([cedula, nombre, cargo, telefono, direccion, email]):
                cell_font_size = 7 if i != len(x_coords) - 1 else 5.5 # Tamaño de letra para la última columna
                self.set_font('Arial', '', cell_font_size)
                self.cell(32, 11, cell, 1, 0, 'C', False, 0)
                self.set_xy(x_coords[i], self.get_y())
                    
            self.ln()
            self.set_xy(10, self.get_y())
        
    # Método para crear el footer
    def footer(self):
        # Agregamos el texto del footer en gris
        self.set_y(-15)
        self.set_font('Arial', '', 8)
        self.set_text_color(127, 127, 127)
        self.cell(0, 10, 'Procol del Valle © Astrocode | Página ' + str(self.page_no()), 0, 0, 'C')

def verifyDirectories(directory, file):
    #Crea la carpeta donde se almacenan los datos
    directory_path = os.path.join(os.path.join(os.environ['USERPROFILE']), directory)
    procol_path = os.path.join(directory_path, 'Procol del Valle')

    if not os.path.exists(procol_path):
        os.makedirs(procol_path)
        excels_path = os.path.join(procol_path, 'Excels')
        pdf_path = os.path.join(procol_path, 'PDF')
        os.makedirs(excels_path)
        os.makedirs(pdf_path)

    if file == "Excel":
        excels_path = os.path.join(procol_path, 'Excels')

        if not os.path.exists(excels_path):
            os.makedirs(excels_path)

        return excels_path
    elif file == "Pdf":
        pdfs_path = os.path.join(procol_path, 'PDF')

        if not os.path.exists(pdfs_path):
            os.makedirs(pdfs_path)
        
        return pdfs_path